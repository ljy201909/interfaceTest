# coding:utf-8
__author__ = 'ljy'
#create_time = 2019.11.05

import openpyxl
import os


class writeExcel():
    dir = 'testData'
    # excel_dir = os.path.dirname(os.getcwd()) + "\\" + dir
    excel_dir = os.getcwd() + "\\" + dir
    # print(excel_dir)
    wb = openpyxl.load_workbook(excel_dir + '\\' + 'data.xlsx')
    sheetnames = wb.sheetnames

    def writeResult(self,id,real,result):
        #打开一个Excel
        # wb = openpyxl.Workbook()
        # sheetnames = wb.sheetnames
        # print(sheetnames)
        # #打开名字为assertSheet的sheet
        ws = self.wb['assertSheet']
        #写入数据，id是行和编号，real是实际返回值，result是结果
        ws.cell(id+1,3).value = real
        ws.cell(id+1,4).value = result
        #保存
        self.wb.save(self.excel_dir + '\\' + 'data.xlsx')


# wr = writeExcel()
# wr.writeResult(1,'0','success')
